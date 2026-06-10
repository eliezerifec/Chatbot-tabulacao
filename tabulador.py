"""
tabulador.py  —  Motor de tabulação universal  —  IFec RJ
==========================================================
Replica a lógica do TabIFec (R) em Python.

Comportamento para perguntas com "Outro. Qual?" codificado:
  - Opções fechadas + Outro ficam numa tabela só
  - "Outro" aparece em fonte normal com total de quem marcou
  - Categorias codificadas aparecem recuadas, em itálico, abaixo de Outro
"""

from __future__ import annotations
import re
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────────────────────────────────────
# TIPOS
# ─────────────────────────────────────────────────────────────────────────────

TIPOS_LABEL = {
    "RU":     "Resposta Única",
    "RM":     "Resposta Múltipla",
    "GRID":   "Matriz / Ranking",
    "ABERTA": "Aberta / Codificada",
    "MEDIA":  "Numérica (Média)",
    "NPS":    "NPS (0-10)",
    "IGNORAR":"Ignorar",
}


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA E set_header
# ─────────────────────────────────────────────────────────────────────────────

def set_header(
    df_raw: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, str], dict[str, str]]:
    """
    Replica set_header() do TabIFec: combina linha 0 + linha 1.

    Retorna (df, tipos_sm, q0_map) onde:
      tipos_sm  → {col: "RU"|"RM"|"ABERTA"} — tipo da linha 1 do SurveyMonkey
      q0_map    → {col: texto_original_linha0} — chave de agrupamento confiável:
                  todas as opções de uma mesma pergunta RM compartilham o mesmo
                  texto da linha 0 (forward-fill), independente de terem '(' no nome.
    """
    row0, row1 = df_raw.iloc[0], df_raw.iloc[1]
    filled, last = [], ""
    for v in row0:
        if pd.notna(v) and str(v).strip() not in ("", "nan"):
            last = str(v).strip()
        filled.append(last)

    cols: list[str] = []
    raw_row1: list[str] = []
    q0_filled: list[str] = []   # texto original da linha 0 (antes de combinar)
    for q, r in zip(filled, row1):
        r_str = str(r).strip() if pd.notna(r) else ""
        raw_row1.append(r_str)
        q0_filled.append(q)
        combined = q + r_str
        combined = re.sub(r"\s*(Response$|Open-Ended Response$|NA$)\s*$", "", combined)
        combined = re.sub(r"\s*\[.*?\]", "", combined)
        cols.append(combined.strip())

    df = df_raw.iloc[2:].copy().reset_index(drop=True)
    seen: dict[str, int] = {}
    unique_cols: list[str] = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            unique_cols.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            unique_cols.append(c)
    df.columns = unique_cols

    tipos_sm: dict[str, str] = {}
    q0_map:   dict[str, str] = {}
    for col_name, r_str, q0_text in zip(unique_cols, raw_row1, q0_filled):
        # Tipo a partir da linha 1
        if r_str == "Response":
            tipos_sm[col_name] = "RU"
        elif r_str == "Open-Ended Response" or r_str.endswith(":"):
            tipos_sm[col_name] = "ABERTA"
        else:
            tipos_sm[col_name] = "RM"
        # Chave de agrupamento = texto da linha 0 (sem sufixo)
        q0_map[col_name] = q0_text.strip()

    return df, tipos_sm, q0_map


def carregar_base(caminho: str) -> pd.DataFrame:
    """Compatibilidade retroativa — retorna só o DataFrame."""
    xl = pd.ExcelFile(caminho)
    df_raw = pd.read_excel(caminho, header=None, sheet_name=xl.sheet_names[0])
    df, _, _ = set_header(df_raw)
    return df


def carregar_base_com_tipos(
    caminho: str,
) -> tuple[pd.DataFrame, dict[str, str], dict[str, str]]:
    """Retorna (df, tipos_sm, q0_map) usando os metadados do SurveyMonkey."""
    xl = pd.ExcelFile(caminho)
    df_raw = pd.read_excel(caminho, header=None, sheet_name=xl.sheet_names[0])
    return set_header(df_raw)


# ─────────────────────────────────────────────────────────────────────────────
# PADRÕES
# ─────────────────────────────────────────────────────────────────────────────

_IGNORAR = {
    "respondent_id","collector_id","date_created","date_modified",
    "ip_address","email_address","first_name","last_name","custom_1",
    "Pesquisador:","Pesquisador","Entrevistador:","Entrevistador",
    "Comentário:","Comentário","Observações:","Observações",
    "Setor:","Setor","TOTAL","NPS","FX_IDADE",
}

_PAD_OUTRO_COL = re.compile(r"\?.*Outr[oa]|Outr[oa].*(\?|:|\.)\s*$", re.IGNORECASE)
_PAD_NPS       = re.compile(r"de 0 a 10|numa escala de 0|probabilidade.*recomendar", re.IGNORECASE)
_PAD_MEDIA     = re.compile(r"de 1 a \d|de \d a \d|quantas? (anos?|vezes?|pessoas?)|qual.*idade|renda.*r\$", re.IGNORECASE)
_PAD_OUTRO_VAL = re.compile(r"Outr[oa].*[:?.]", re.IGNORECASE)


def _e_outro(col: str) -> bool:
    return bool(_PAD_OUTRO_COL.search(col))


def _e_cod(col: str) -> bool:
    return col.endswith("_cod")


def _detectar_grid(df: pd.DataFrame, cols: list[str]) -> bool:
    """
    Detecta perguntas do tipo GRID / Ranking:
    - Múltiplas colunas
    - Cada coluna tem poucos valores únicos (≤ 10)
    - Os valores se repetem em ≥ 50 % das colunas (mesma escala)
    Exemplos: Sim/Talvez/Não por item, classificação 1-5 por critério.
    """
    if len(cols) <= 1:
        return False
    per_col: list[set] = []
    all_vals: set = set()
    for c in cols:
        vals = set(_col(df, c).dropna().astype(str).str.strip())
        vals -= {"", "nan", "-"}
        if not vals:
            continue
        per_col.append(vals)
        all_vals |= vals
    if not per_col or len(all_vals) == 0 or len(all_vals) > 10:
        return False
    intersection = per_col[0].copy()
    for v in per_col[1:]:
        intersection &= v
    return len(intersection) >= 2 and (len(intersection) / len(all_vals)) >= 0.5


# ─────────────────────────────────────────────────────────────────────────────
# DETECÇÃO DE PERGUNTAS
# ─────────────────────────────────────────────────────────────────────────────

def _prefixo_pergunta(col: str) -> str:
    """
    Extrai o prefixo que identifica a pergunta-raiz de uma coluna.

    Ordem de prioridade:
      1. '?' → corta aqui (ex: "Qual curso?Cabeleireiro" → "Qual curso?")
      2. '…' (U+2026) seguido de letra maiúscula ou dígito → separador
         entre enunciado e opção em bases já processadas do SurveyMonkey
         (ex: "A escola é…Pública (do governo...)" → "A escola é…")
      3. '...' (3 pontos) idem
      4. ':' no final da string → pergunta autônoma terminada em dois-pontos
      5. Qualquer outro caso → coluna autônoma (NÃO divide em ' (')
    """
    # 1. '?'
    idx_q = col.find('?')
    if idx_q != -1:
        return col[:idx_q + 1]

    # 2/3. Reticências ('…' U+2026 ou '...')
    for ell in ('…', '...'):
        idx_e = col.find(ell)
        if idx_e != -1:
            after = col[idx_e + len(ell):]
            # Só corta se o que segue parece ser uma opção (maiúscula, dígito ou '(')
            if after and (after[0].isupper() or after[0].isdigit() or after[0] == '('):
                return col[:idx_e + len(ell)]

    # 4. ':' no final → pergunta autônoma
    if col.rstrip().endswith(':'):
        return col

    # 5. Sem separador claro → coluna autônoma (não divide em ' (')
    return col


def detectar_perguntas(
    df: pd.DataFrame,
    tipos_sm: dict[str, str] | None = None,
    q0_map:   dict[str, str] | None = None,
) -> list[dict]:
    """
    Agrupa colunas por pergunta-raiz.

    Quando q0_map (linha 0 do SurveyMonkey) estiver disponível, usa o texto
    original da pergunta como chave de agrupamento — resolve o problema de RM
    onde as opções têm '(' no nome, enganando o _prefixo_pergunta.
    Também detecta automaticamente a coluna "Outro. Qual?" pelo tipos_sm.

    Sem q0_map, usa heurística de prefixo (até '?' ou ' (').
    """
    todas = list(df.columns)

    princ = [c for c in todas
             if c not in _IGNORAR
             and not c.startswith("Original_")
             and not _e_cod(c)]

    # ── Passo 1: chave de agrupamento por coluna ──────────────────────────────
    # Com q0_map → usa texto linha 0 (confiável).
    # Sem q0_map → usa _prefixo_pergunta (heurística).
    prefixo_map:   dict[str, str]        = {}
    grupos_prefixo: dict[str, list[str]] = {}

    for col in princ:
        if q0_map:
            pref = q0_map.get(col, "").strip() or _prefixo_pergunta(col)
        else:
            pref = _prefixo_pergunta(col)
        prefixo_map[col] = pref
        grupos_prefixo.setdefault(pref, []).append(col)

    # ── Passo 2: montar perguntas na ordem em que aparecem no DataFrame ────────
    perguntas, num = [], 1
    vistas_prefixos: set[str] = set()

    for col in todas:
        if col not in prefixo_map:
            continue
        pref = prefixo_map[col]
        if pref in vistas_prefixos:
            continue
        if col in _IGNORAR or col.startswith("Original_"):
            continue
        vistas_prefixos.add(pref)

        grupo = grupos_prefixo[pref]

        # ── Separar colunas de opção vs. campo-aberto "Outro" ─────────────────
        # Com tipos_sm: campo aberto detectado via "ABERTA" (linha 1 = "Open-Ended Response").
        # Sem tipos_sm: usa regex _PAD_OUTRO_COL no nome da coluna.
        if tipos_sm:
            # "Outro(s). Qual?" nem sempre vem como "Open-Ended Response" na
            # linha 1 do SurveyMonkey — detecta também pelo nome da coluna.
            cols_outros = [c for c in grupo
                           if tipos_sm.get(c) == "ABERTA" or _e_outro(c)]
            cols_princ  = [c for c in grupo if c not in cols_outros]
        else:
            cols_outros = [c for c in grupo if _e_outro(c)]
            cols_princ  = [c for c in grupo if c not in cols_outros]

        if not cols_princ:
            # Pergunta 100% aberta (ex.: subcampos "1:", "2:", "3:") —
            # tabula as respostas codificadas (_cod) de todas as colunas.
            if not cols_outros:
                continue
            if q0_map:
                nome_pergunta = (q0_map.get(cols_outros[0], pref) or pref).strip()
            else:
                nome_pergunta = pref
            if nome_pergunta in _IGNORAR:
                continue
            cols_cod = [c + "_cod" for c in cols_outros if c + "_cod" in df.columns]
            perguntas.append({
                "num":         f"P{num:02d}",
                "pergunta":    nome_pergunta,
                "tipo":        "ABERTA",
                "colunas":     cols_outros,
                "cols_outros": cols_outros,
                "col_cod":     cols_cod[0] if cols_cod else None,
                "cols_cod":    cols_cod,
                "nota":        _nota_padrao("ABERTA"),
                "ativo":       True,
            })
            num += 1
            continue

        # Procurar coluna _cod associada
        # Prioridade 1: coluna com sufixo _cod (base raw + codificação Python)
        col_cod = None
        for c in cols_outros:
            cand = c + "_cod"
            if cand in df.columns:
                col_cod = cand
                break

        # Prioridade 2: busca por prefixo entre colunas _cod existentes
        if col_cod is None:
            pref_curto = pref[:25]
            for c in todas:
                if _e_cod(c) and c[:-4].startswith(pref_curto):
                    col_cod = c
                    break

        # Prioridade 3 (base processada — equivale ao tabela_o2 do R com serie=T):
        # a própria coluna "Outro" já contém as categorias codificadas (sem sufixo _cod).
        # Igual ao `select(matches("\\?.*Outro.*(\\?|:)"))` → `separate(sep=", ")`
        if col_cod is None and cols_outros:
            col_cod = cols_outros[0]

        # Nome da pergunta:
        #   - Com q0_map e grupo RM: texto da linha 0 (pergunta sem sufixo de opção)
        #   - Sem q0_map ou grupo único: usa prefixo ou col_name inteiro
        if q0_map and len(grupo) > 1 and cols_princ:
            nome_pergunta = (q0_map.get(cols_princ[0], pref) or pref).strip()
        elif len(grupo) > 1:
            nome_pergunta = pref
        else:
            nome_pergunta = cols_princ[0]

        tipo = _detectar_tipo(df, nome_pergunta, cols_princ, tipos_sm or {})
        if tipo == "IGNORAR":
            continue

        # Forçar RM quando o grupo tem mais de uma coluna principal
        # (não aplica para GRID, NPS e ABERTA que já têm lógica própria)
        if tipo not in ("NPS", "ABERTA", "GRID") and len(cols_princ) > 1:
            tipo = "RM"

        # Nota de "Não soube avaliar"
        nota_nsa = ""
        if tipo in ("RU", "RM"):
            n_nsa = 0
            for cp in cols_princ:
                serie_cp = _col(df, cp).astype(str).str.strip()
                n_nsa += int(serie_cp.str.lower().str.contains(
                    r"não soube avaliar|nao soube avaliar", regex=True).sum())
            if n_nsa > 0:
                nota_nsa = (f'{n_nsa} pessoa{"(s)" if n_nsa > 1 else ""} '
                            f'respondeu/responderam "Não soube avaliar".')

        nota = _nota_padrao(tipo)
        if nota_nsa:
            nota = (nota + "\n" + nota_nsa) if nota else nota_nsa

        perguntas.append({
            "num":         f"P{num:02d}",
            "pergunta":    nome_pergunta,
            "tipo":        tipo,
            "colunas":     cols_princ,
            "cols_outros": cols_outros,
            "col_cod":     col_cod,
            "nota":        nota,
            "ativo":       True,
        })
        num += 1

    return perguntas


def _detectar_tipo(df: pd.DataFrame, raiz: str, cols: list[str],
                   tipos_sm: dict[str, str] | None = None) -> str:
    raiz_l = raiz.lower()
    if raiz in _IGNORAR or raiz.startswith("Original_") or not raiz.strip():
        return "IGNORAR"

    # NPS e MEDIA têm prioridade — detecção por padrão textual, independente do SM
    if _PAD_NPS.search(raiz_l):
        return "NPS"

    cols_s = [c for c in cols if not _e_outro(c)]

    # MEDIA só vale para pergunta de coluna única; grupos multi-coluna com
    # escala numérica (ex.: "Classifique de 1 a 5" por item) são GRID.
    if _PAD_MEDIA.search(raiz_l) and len(cols_s) == 1:
        sample = pd.to_numeric(_col(df, cols_s[0]), errors="coerce").dropna()
        if len(sample) > 0 and sample.mean() > 0:
            return "MEDIA"

    # Usa o tipo da linha 1 do SurveyMonkey como fonte primária
    if tipos_sm and cols_s:
        tipo_sm = tipos_sm.get(cols_s[0])
        if tipo_sm in ("RU", "ABERTA", "RM"):
            # RM com valores de escala → GRID
            if tipo_sm == "RM" and _detectar_grid(df, cols_s):
                return "GRID"
            return tipo_sm

    # Fallback heurístico (para bases sem metadado SM ou com tipos ausentes)
    if len(cols_s) > 1:
        if _detectar_grid(df, cols_s):
            return "GRID"
        if sum(1 for c in cols_s if _col(df, c).notna().any()) > 1:
            return "RM"

    if not cols_s:
        return "IGNORAR"

    serie = _col(df, cols_s[0]).dropna().astype(str).pipe(lambda s: s[s.str.strip() != ""])
    if len(serie) == 0:
        return "IGNORAR"
    if serie.nunique() / len(serie) > 0.35:
        return "ABERTA"
    return "RU"


def _auto_ordenar(labels: list[str]) -> list[str]:
    """
    Tenta ordenar labels de faixas de preço / salário pelo valor inicial.
    Retorna a lista original se não reconhecer o padrão.
    """
    _RE_NENHUM = re.compile(r"não tenho|nao tenho|sem condição|nao tenho", re.I)
    _RE_ACIMA  = re.compile(r"acima|mais de", re.I)
    _RE_NUM    = re.compile(r"[\d.,]+")

    def _key(lbl: str):
        if _RE_NENHUM.search(lbl):
            return -1.0
        nums = _RE_NUM.findall(lbl.replace(".", "").replace(",", "."))
        if not nums:
            return float("inf")
        val = float(nums[0])
        if _RE_ACIMA.search(lbl):
            val += 1e9     # acima → vai para o final
        return val

    try:
        sorted_labels = sorted(labels, key=_key)
        # Só aplica se parecer com faixas monetárias/salariais
        tem_faixa = any(
            re.search(r"R\$|salário|salario|mínimo|minimo", l, re.I)
            for l in labels
        )
        return sorted_labels if tem_faixa else labels
    except Exception:
        return labels


def _nota_padrao(tipo: str) -> str:
    return {
        "RU":     "Pergunta com resposta única.",
        "RM":     "Pergunta com resposta múltipla.",
        "ABERTA": "Pergunta aberta (espontânea).",
        "MEDIA":  "Média das respostas.",
        "NPS":    "Índice de Recomendação (NPS).",
    }.get(tipo, "")


def _col(df: pd.DataFrame, nome: str) -> pd.Series:
    """Retorna sempre uma pd.Series, mesmo que a coluna seja duplicada."""
    resultado = df[nome]
    if isinstance(resultado, pd.DataFrame):
        return resultado.iloc[:, 0]
    return resultado


def _cols(df: pd.DataFrame, nomes: list) -> pd.DataFrame:
    """Retorna DataFrame com as colunas pedidas, desduplicando cada uma."""
    partes = [_col(df, n).rename(n) for n in nomes]
    return pd.concat(partes, axis=1)


# ─────────────────────────────────────────────────────────────────────────────
# TABULAÇÃO
# Retorna DataFrame com colunas: [" ", "Total", "%", "is_sub"]
# is_sub=True → linha em itálico recuado no Excel (sub-item de Outro)
# ─────────────────────────────────────────────────────────────────────────────

def tabular_ru_rm(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    cols_p   = pergunta["colunas"]
    cols_o   = pergunta.get("cols_outros", [])
    col_cod  = pergunta.get("col_cod")

    # ── Opções fechadas ───────────────────────────────────────────────────────
    df_long = (_cols(df, cols_p).stack().reset_index(drop=True).to_frame(name=" "))
    df_long[" "] = df_long[" "].astype(str).str.strip()
    df_long = df_long[
        df_long[" "].notna() &
        ~df_long[" "].isin(["", "-", "nan"]) &
        ~df_long[" "].str.startswith("NÃO SE APLICA") &
        ~df_long[" "].str.match(_PAD_OUTRO_VAL) &
        # Exclui o valor "Outro"/"Outra" da opção checkbox — contado separadamente
        ~df_long[" "].str.strip().str.lower().isin(["outro", "outra"]) &
        ~df_long[" "].str.lower().str.contains(r"não soube avaliar|nao soube avaliar", regex=True)
    ]

    freq = df_long[" "].value_counts().reset_index()
    freq.columns = [" ", "Total"]
    freq["is_sub"] = False

    # ── Contagem de Outro ─────────────────────────────────────────────────────
    n_outro = 0
    if cols_o:
        n_outro = int(
            _cols(df, cols_o).apply(
                lambda col: col.notna() & ~col.astype(str).str.strip().isin(["", "-", "nan"]),
                axis=0
            ).any(axis=1).sum()
        )
    elif col_cod and col_cod in df.columns:
        serie_outro = _col(df, col_cod)
        n_outro = int(
            (serie_outro.notna() &
             ~serie_outro.astype(str).str.strip().isin(["", "-", "nan"])).sum()
        )

    # ── Sub-itens codificados ─────────────────────────────────────────────────
    sub_df = pd.DataFrame(columns=[" ", "Total", "is_sub"])
    if col_cod and col_cod in df.columns:
        serie_cod = (
            _col(df, col_cod).dropna().astype(str)
            .pipe(lambda s: s[~s.str.strip().isin(["", "-", "nan"])])
        )
        if len(serie_cod) > 0:
            if not cols_o:
                n_outro = len(serie_cod)
            expandido = serie_cod.str.split(", ").explode().str.strip()
            expandido = expandido[expandido != ""]
            sf = expandido.value_counts().reset_index()
            sf.columns = [" ", "Total"]
            sf["is_sub"] = True
            sub_df = sf

    # ── Ordenação das opções PRINCIPAIS (antes de inserir Outro) ─────────────
    # 1) Ordem manual: excluímos "Outro" pois sua posição é sempre ao final
    # 2) Ordem automática de faixas de preço/salário
    ordem_manual = [str(o).strip() for o in (pergunta.get("ordem") or [])
                    if str(o).strip() and str(o).strip().lower() != "outro"]
    main_lbls = list(freq[" "])
    ref = ordem_manual if ordem_manual else _auto_ordenar(main_lbls)
    if ref != main_lbls:
        ordered = [freq[freq[" "] == lbl] for lbl in ref if lbl in main_lbls]
        remaining = freq[~freq[" "].isin(set(ref))]
        freq = pd.concat(ordered + [remaining], ignore_index=True)

    # ── Montagem final: "Outro" + sub-itens abaixo (ou ocultos) ─────────────
    mostrar_outro = pergunta.get("mostrar_outro", True)

    if n_outro > 0 or not sub_df.empty:
        outro_row = pd.DataFrame({" ": ["Outro"], "Total": [n_outro], "is_sub": [False]})
        if mostrar_outro and not sub_df.empty:
            # Sub-itens codificados aparecem ABAIXO do "Outro"
            freq = pd.concat([freq, outro_row, sub_df], ignore_index=True)
        else:
            # Sem detalhamento: apenas a linha "Outro"
            freq = pd.concat([freq, outro_row], ignore_index=True)

    # ── Base (respondentes únicos) ────────────────────────────────────────────
    todas_cols = cols_p + cols_o
    respondentes = int(_cols(df, todas_cols).apply(
        lambda row: any(
            pd.notna(v) and str(v).strip() not in ("", "-", "NÃO SE APLICA")
            for v in row
        ), axis=1
    ).sum()) or 1

    total = pd.DataFrame({" ": ["Total"], "Total": [respondentes], "is_sub": [False]})
    freq  = pd.concat([freq, total], ignore_index=True)

    freq["%"] = freq.apply(
        lambda r: (r["Total"] / respondentes)
        if isinstance(r["Total"], (int, np.integer)) and r[" "] != "Total"
        else (1.0 if r[" "] == "Total" else "-"),
        axis=1
    )

    return freq[[" ", "Total", "%", "is_sub"]]


def tabular_aberta(df: pd.DataFrame, pergunta: dict, sep: str = ", ") -> pd.DataFrame:
    # Perguntas abertas com vários subcampos ("1:", "2:", "3:") têm uma coluna
    # codificada por subcampo — soma as menções de todas.
    cols = [c for c in (pergunta.get("cols_cod") or []) if c in df.columns]
    if not cols:
        col_cod = pergunta.get("col_cod")
        cols = [col_cod if (col_cod and col_cod in df.columns)
                else pergunta["colunas"][0]]

    series = pd.concat([_col(df, c) for c in cols], ignore_index=True)
    serie = (series.dropna().astype(str)
             .pipe(lambda s: s[~s.str.strip().isin(["", "-", "nan"])]))

    expandido = serie.str.split(sep).explode().str.strip()
    expandido = expandido[expandido != ""]
    expandido = expandido[~expandido.str.match(_PAD_OUTRO_VAL)]

    freq = expandido.value_counts().reset_index()
    freq.columns = [" ", "Total"]
    freq["is_sub"] = False

    # Respondentes = linhas com pelo menos uma resposta em algum subcampo
    respondida = pd.Series(False, index=df.index)
    for c in cols:
        s = _col(df, c)
        respondida = respondida | (s.notna() & ~s.astype(str).str.strip()
                                   .isin(["", "-", "nan"]))
    respondentes = int(respondida.sum()) or 1
    total = pd.DataFrame({" ": ["Total"], "Total": [respondentes], "is_sub": [False]})
    freq  = pd.concat([freq, total], ignore_index=True)
    freq["%"] = freq["Total"].apply(
        lambda x: x / respondentes if isinstance(x, (int, np.integer)) else "-"
    )
    return freq[[" ", "Total", "%", "is_sub"]]


def tabular_media(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    valores = pd.to_numeric(_col(df, pergunta["colunas"][0]), errors="coerce").dropna()
    media   = round(float(valores.mean()), 2) if len(valores) > 0 else 0
    rows = [
        [" ", "Total", "%",  "is_sub"],
        ["Média",  media,        "-", False],
        ["Total",  len(valores), "-", False],
    ]
    return pd.DataFrame(rows[1:], columns=rows[0])


def tabular_nps(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    """
    Retorna 4 blocos separados por linha vazia, como no padrão:
      1. Distribuição de notas (0-10)
      2. Grupos (Promotores / Neutros / Detratores)
      3. NPS
      4. Média
    Cada bloco tem seu próprio cabeçalho e linha Total.
    Usamos is_sub=False para tudo; o exportar_excel os escreve em sequência.
    Sinalizamos a separação com linhas sentinela is_sep=True.
    """
    valores = pd.to_numeric(_col(df, pergunta["colunas"][0]), errors="coerce").dropna()
    total   = len(valores)
    if total == 0:
        return pd.DataFrame(columns=[" ", "Total", "%", "is_sub", "is_sep"])

    prom = int((valores >= 9).sum())
    neut = int(((valores >= 7) & (valores <= 8)).sum())
    detr = int((valores <= 6).sum())
    nps  = round((prom - detr) / total * 100, 1)
    media = round(float(valores.mean()), 6)

    rows = []

    # Bloco 1 — distribuição de notas (ordem decrescente)
    for nota_val, cnt in valores.value_counts().sort_index(ascending=False).items():
        rows.append((int(nota_val), int(cnt), int(cnt)/total, False, False))
    rows.append(("Total", total, 1.0, False, False))

    # Separador
    rows.append((None, None, None, False, True))

    # Bloco 2 — grupos
    rows.append(("Promotores",  prom,  prom/total,  False, False))
    rows.append(("Neutros",     neut,  neut/total,  False, False))
    rows.append(("Detratores",  detr,  detr/total,  False, False))
    rows.append(("Total",       total, 1.0,         False, False))

    # Separador
    rows.append((None, None, None, False, True))

    # Bloco 3 — NPS
    rows.append((f"NPS", round(nps, 5), "-", False, False))

    # Separador
    rows.append((None, None, None, False, True))

    # Bloco 4 — Média
    rows.append(("Média", media, "-", False, False))
    rows.append(("Total", total, "-", False, False))

    return pd.DataFrame(rows, columns=[" ", "Total", "%", "is_sub", "is_sep"])


def tabular_grid_item(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """
    Tabula um item de pergunta GRID (uma coluna com escala, ex: 1-5 ou Sim/Não).
    Ordena numericamente quando possível.
    """
    serie = (_col(df, col).dropna().astype(str)
             .pipe(lambda s: s[~s.str.strip().isin(["", "nan", "-"])]))
    freq = serie.value_counts().reset_index()
    freq.columns = [" ", "Total"]
    # Ordena numericamente ou alfabeticamente
    try:
        freq[" "] = pd.to_numeric(freq[" "])
        freq = freq.sort_values(" ").reset_index(drop=True)
        freq[" "] = freq[" "].astype(str)
    except Exception:
        freq = freq.sort_values(" ").reset_index(drop=True)
    freq["is_sub"] = False
    freq["is_sep"] = False
    respondentes = len(serie) or 1
    total = pd.DataFrame({
        " ": ["Total"], "Total": [respondentes],
        "is_sub": [False], "is_sep": [False]
    })
    freq = pd.concat([freq, total], ignore_index=True)
    freq["%"] = freq.apply(
        lambda r: r["Total"] / respondentes if r[" "] != "Total" else 1.0, axis=1
    )

    # Garante valores inteiros para escalas numéricas ("1.0" → "1")
    def _to_int_str(v):
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else v
        except Exception:
            return v
    freq[" "] = freq[" "].apply(
        lambda v: _to_int_str(v) if v not in ("Total",) else v
    )

    return freq[[" ", "Total", "%", "is_sub", "is_sep"]]


def preparar_aberturas(
    df: pd.DataFrame,
    cols_abertura: list[str]
) -> list[tuple[str, pd.DataFrame]]:
    """
    Prepara a lista de (label, df_subset) para cada abertura (cruzamento).
    Sempre começa com ("Total", df) para o total geral.

    cols_abertura: nomes das colunas do df para cruzar.
                   Para cada coluna, um par por valor único é adicionado.
    """
    result: list[tuple[str, pd.DataFrame]] = [("Total", df)]
    for col in (cols_abertura or []):
        if col not in df.columns:
            continue
        vals = sorted(df[col].dropna().astype(str).unique())
        for val in vals:
            df_sub = df[df[col].astype(str) == val].reset_index(drop=True)
            if len(df_sub) > 0:
                result.append((val, df_sub))
    return result


def tabular_pergunta(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    t = pergunta["tipo"]
    if t in ("RU", "RM"): return tabular_ru_rm(df, pergunta)
    if t == "ABERTA":     return tabular_aberta(df, pergunta)
    if t == "MEDIA":      return tabular_media(df, pergunta)
    if t == "NPS":        return tabular_nps(df, pergunta)
    if t == "GRID":
        # Fallback: tabula primeira coluna como RU simples
        if pergunta.get("colunas"):
            return tabular_grid_item(df, pergunta["colunas"][0])
    return pd.DataFrame(columns=[" ", "Total", "%", "is_sub"])


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAÇÃO EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _grid_item_label(pergunta: dict, col: str) -> str:
    """Extrai o nome do item GRID (sufixo após o prefixo da pergunta)."""
    pref = pergunta.get("pergunta", "")
    if col.startswith(pref):
        label = col[len(pref):].lstrip("?").strip()
    else:
        label = col
    return label or col


def _ab_col(i: int) -> tuple[int, int]:
    """
    Retorna (col_count, col_pct) para a i-ésima abertura.
    i=0 → Total (colunas B=2, C=3)
    i≥1 → separador em D=4, abertura em E=5,F=6 / G=7,H=8 / ...
    """
    if i == 0:
        return 2, 3
    return 4 + (i - 1) * 2 + 1, 4 + (i - 1) * 2 + 2


def exportar_excel(df: pd.DataFrame, perguntas: list[dict],
                   saida: str, titulo: str = "Pesquisa",
                   total_respostas: int = None,
                   metodologia: str = None, rodape: str = None,
                   aberturas_cols: list[str] | None = None,
                   filtro_col: str | None = None,
                   filtro_cols: list[str] | None = None):

    from openpyxl.styles import Border, Side

    if total_respostas is None:
        total_respostas = len(df)

    _MET = metodologia or (
        "Metodologia: Pesquisa do tipo quantitativa com coleta de dados "
        "através de questionário (com perguntas abertas e fechadas) "
        "aplicado de forma presencial."
    )
    _ROD = rodape or (
        'Todas as perguntas foram obrigatórias. Nas perguntas com múltiplas respostas, '
        'os respondentes podem apontar mais de uma opção, por isso a soma das frequências '
        'passa de 100%. Nas perguntas com campo aberto na opção "Outro", as frequências '
        'apresentadas são das respostas válidas categorizadas. Os respondentes também podem '
        'apresentar mais de uma resposta no campo aberto "Outro".'
    )

    # ── Paleta ───────────────────────────────────────────────────────────────
    COR_TITULO   = "984806"
    COR_DATA     = "E36C0A"
    COR_PERGUNTA = "984806"
    COR_TEXTO    = "262626"
    COR_NOTA     = "404040"
    COR_HDR_BG   = "E36C0A"
    COR_HDR_FG   = "FFFFFF"

    FILL_HDR = PatternFill("solid", fgColor=COR_HDR_BG)
    AL_WRAP  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_IND   = Alignment(horizontal="left",   vertical="center", wrap_text=True, indent=2)
    BD_DOUBLE = Border(bottom=Side(border_style="double"))
    BD_MEDIUM = Border(top=Side(border_style="medium"),
                       bottom=Side(border_style="medium"))

    def F(bold=False, size=12, color=COR_TEXTO, italic=False, name="Cambria"):
        return Font(name=name, bold=bold, size=size, color=color, italic=italic)

    def num(v):
        if isinstance(v, (pd.Series, pd.DataFrame)):
            v = v.iloc[0] if len(v) > 0 else 0
        try:
            fv = float(v)
            return int(fv) if fv == int(fv) else round(fv, 2)
        except Exception:
            return v

    # ── Aberturas ─────────────────────────────────────────────────────────────
    ab_list = preparar_aberturas(df, aberturas_cols or [])
    n_ab    = len(ab_list)  # inclui Total

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    def _escrever_sheet(ws: object, df_ws: pd.DataFrame,
                        ab_ws: list, total_ws: int) -> None:
        """Preenche uma aba do workbook com cabeçalho + perguntas."""
        nonlocal FILL_HDR, AL_WRAP, AL_CTR, AL_LEFT, AL_IND
        nonlocal BD_DOUBLE, BD_MEDIUM

        n_ab_ws = len(ab_ws)
        has_sep = n_ab_ws > 1  # separador entre Total e aberturas

        # ── Helpers de posição de coluna ──────────────────────────────────────
        def col_count(i: int) -> int:
            """Coluna para contagem da i-ésima abertura (0=Total)."""
            return 2 if i == 0 else (5 + (i - 1) * 2)

        def col_pct(i: int) -> int:
            """Coluna para % da i-ésima abertura (0=Total)."""
            return 3 if i == 0 else (6 + (i - 1) * 2)

        # ── Linha de cabeçalho de coluna ──────────────────────────────────────
        def _write_col_headers(row_n: int, label_col_a: str = "") -> None:
            c = ws.cell(row=row_n, column=1, value=label_col_a or None)
            c.font = F(bold=True, size=12, color=COR_HDR_FG, name="Cambria")
            c.fill = FILL_HDR
            c.alignment = AL_CTR if not label_col_a else AL_LEFT
            c.border = BD_DOUBLE

            ws.cell(row=row_n, column=col_count(0),
                    value="Total").font = F(bold=True, size=12, color=COR_HDR_FG, name="Cambria")
            ws.cell(row=row_n, column=col_count(0)).fill = FILL_HDR
            ws.cell(row=row_n, column=col_count(0)).alignment = AL_CTR
            ws.cell(row=row_n, column=col_count(0)).border = BD_DOUBLE

            ws.cell(row=row_n, column=col_pct(0),
                    value="%").font = F(bold=True, size=12, color=COR_HDR_FG, name="Cambria")
            ws.cell(row=row_n, column=col_pct(0)).fill = FILL_HDR
            ws.cell(row=row_n, column=col_pct(0)).alignment = AL_CTR
            ws.cell(row=row_n, column=col_pct(0)).number_format = "0.0%"
            ws.cell(row=row_n, column=col_pct(0)).border = BD_DOUBLE

            if has_sep:
                ws.cell(row=row_n, column=4, value=None)  # separador
                for i, (ab_label, _) in enumerate(ab_ws[1:], 1):
                    cc = col_count(i)
                    cp = col_pct(i)
                    cl = ws.cell(row=row_n, column=cc, value=ab_label)
                    cl.font = F(bold=True, size=10, color=COR_HDR_FG, name="Cambria")
                    cl.fill = FILL_HDR
                    cl.alignment = AL_CTR
                    cl.border = BD_DOUBLE
                    cp_c = ws.cell(row=row_n, column=cp, value="%")
                    cp_c.font = F(bold=True, size=10, color=COR_HDR_FG, name="Cambria")
                    cp_c.fill = FILL_HDR
                    cp_c.alignment = AL_CTR
                    cp_c.number_format = "0.0%"
                    cp_c.border = BD_DOUBLE
            ws.row_dimensions[row_n].height = 16.5

        # ── Escreve uma linha de dado ──────────────────────────────────────────
        def _write_data_row(row_n: int, label: str, tipo: str,
                            lookups: list[dict],
                            is_sub: bool = False, is_tot: bool = False,
                            is_nps_lbl: bool = False) -> None:
            is_rm_total = is_tot and tipo == "RM"
            if is_sub:
                f_lbl = F(italic=True, size=9, color="555555", name="Calibri")
                a_lbl = AL_IND
                h = 13
            elif is_tot:
                f_lbl = F(size=12, name="Cambria")
                a_lbl = AL_LEFT
                h = 16.5
            else:
                f_lbl = F(size=12, name="Cambria")
                a_lbl = AL_LEFT
                h = 15.75

            c1 = ws.cell(row=row_n, column=1, value=label)
            c1.font = f_lbl
            c1.alignment = a_lbl
            if is_tot:
                c1.border = BD_MEDIUM

            for i, lkup in enumerate(lookups):
                raw_count = lkup.get(label, {}).get("Total", 0 if not is_nps_lbl else "-")
                raw_pct   = lkup.get(label, {}).get("%",     "-")

                f_data = F(italic=is_sub, size=9 if is_sub else 12,
                           color="555555" if is_sub else COR_TEXTO,
                           name="Calibri" if is_sub else "Cambria")

                val_n = num(raw_count) if not is_nps_lbl and raw_count != "-" else raw_count
                c_cnt = ws.cell(row=row_n, column=col_count(i), value=val_n)
                c_cnt.font = f_data
                c_cnt.alignment = AL_CTR
                if is_tot:
                    c_cnt.border = BD_MEDIUM

                # RM: linha Total sempre com "-" em %
                if is_rm_total:
                    val_p, fmt_p = "-", "General"
                elif isinstance(raw_pct, float) and not np.isnan(raw_pct):
                    val_p, fmt_p = raw_pct, "0.0%"
                else:
                    val_p = str(raw_pct) if raw_pct is not None else "-"
                    fmt_p = "General"
                c_pct = ws.cell(row=row_n, column=col_pct(i), value=val_p)
                c_pct.font = F(italic=is_sub, size=9 if is_sub else 12,
                               color="555555" if is_sub else COR_TEXTO,
                               name="Calibri" if is_sub else "Cambria")
                c_pct.alignment = AL_CTR
                c_pct.number_format = fmt_p
                if is_tot:
                    c_pct.border = BD_MEDIUM

            if has_sep:
                ws.cell(row=row_n, column=4, value=None)
            ws.row_dimensions[row_n].height = h

        # ── Construir lookup por label para todas as aberturas ─────────────────
        def _build_lookups(pergunta: dict, df_list: list) -> list[dict]:
            """
            Para cada df em df_list, computa tabular_pergunta e retorna
            dict {label: {"Total": n, "%": p}}.
            """
            result = []
            for _, df_a in df_list:
                tab = tabular_pergunta(df_a, pergunta)
                lkup: dict[str, dict] = {}
                for _, r in tab.iterrows():
                    lkup[str(r[" "])] = {"Total": r["Total"], "%": r["%"]}
                result.append(lkup)
            return result

        def _build_lookups_item(df_list: list, col: str) -> list[dict]:
            result = []
            for _, df_a in df_list:
                tab = tabular_grid_item(df_a, col)
                lkup = {str(r[" "]): {"Total": r["Total"], "%": r["%"]}
                        for _, r in tab.iterrows()}
                result.append(lkup)
            return result

        # ── Cabeçalho da aba ──────────────────────────────────────────────────
        row = 1
        c = ws.cell(row=row, column=1, value=titulo)
        c.font = F(bold=True, size=18, color=COR_TITULO, name="Palatino")
        ws.row_dimensions[row].height = 22.5
        row += 1

        c = ws.cell(row=row, column=1,
                    value=f"Total de respostas obtidas: {total_ws}")
        c.font = F(bold=True, size=12, color=COR_DATA, name="Cambria")
        ws.row_dimensions[row].height = 15.75
        row += 1
        ws.row_dimensions[row].height = 15.75
        row += 1

        c = ws.cell(row=row, column=1, value=_MET)
        c.font = F(bold=True, size=12, color=COR_TITULO, name="Palatino")
        c.alignment = AL_WRAP
        ws.row_dimensions[row].height = 15.75
        row += 1

        c = ws.cell(row=row, column=1, value=_ROD)
        c.font = F(bold=False, size=10, color=COR_NOTA, italic=True, name="Calibri")
        c.alignment = AL_WRAP
        ws.row_dimensions[row].height = 15.75
        row += 1
        ws.row_dimensions[row].height = 15.75
        row += 1

        # ── Perguntas ──────────────────────────────────────────────────────────
        ativas = [p for p in perguntas
                  if p.get("ativo", True) and p["tipo"] != "IGNORAR"]

        for p in ativas:
            tipo = p["tipo"]

            # Título da pergunta
            c = ws.cell(row=row, column=1, value=f"{p['num']}. {p['pergunta']}")
            c.font = F(bold=False, size=12, color=COR_PERGUNTA, italic=True, name="Calibri")
            c.alignment = AL_WRAP
            ws.row_dimensions[row].height = 15.75
            row += 1

            if tipo == "GRID":
                # ── GRID: sub-tabela por coluna ───────────────────────────────
                for col in p.get("colunas", []):
                    item_lbl = _grid_item_label(p, col)
                    lookups  = _build_lookups_item(ab_ws, col)
                    base_tab = tabular_grid_item(df_ws, col)

                    _write_col_headers(row, label_col_a=item_lbl)
                    row += 1

                    for _, dr in base_tab.iterrows():
                        if dr.get("is_sep", False):
                            continue
                        lbl    = str(dr[" "]) if pd.notna(dr[" "]) else ""
                        is_tot = lbl.strip().lower() == "total"
                        _write_data_row(row, lbl, tipo, lookups,
                                        is_sub=bool(dr.get("is_sub", False)),
                                        is_tot=is_tot)
                        row += 1

            else:
                # ── Pergunta normal (RU/RM/ABERTA/MEDIA/NPS) ──────────────────
                base_tab = tabular_pergunta(df_ws, p)
                if base_tab.empty:
                    continue

                lookups = _build_lookups(p, ab_ws)

                _write_col_headers(row)
                row += 1

                for _, dr in base_tab.iterrows():
                    lbl = str(dr[" "]) if pd.notna(dr[" "]) else ""

                    # Separador NPS
                    if dr.get("is_sep", False):
                        ws.row_dimensions[row].height = 8
                        row += 1
                        _write_col_headers(row)
                        row += 1
                        continue

                    is_tot = lbl.strip().lower() == "total"
                    is_nps = lbl in ("NPS",) or lbl.startswith("NPS =")
                    _write_data_row(row, lbl, tipo, lookups,
                                    is_sub=bool(dr.get("is_sub", False)),
                                    is_tot=is_tot, is_nps_lbl=is_nps)
                    row += 1

            # Nota de rodapé
            if p.get("nota"):
                c = ws.cell(row=row, column=1, value=p["nota"])
                c.font = F(bold=False, size=10, color=COR_NOTA, italic=True, name="Calibri")
                c.alignment = AL_WRAP
                ws.row_dimensions[row].height = 15.75
                row += 1

            row += 1  # linha em branco

        # ── Largura das colunas ────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 90.71
        ws.column_dimensions["B"].width = 10.0
        ws.column_dimensions["C"].width = 9.14
        if has_sep:
            ws.column_dimensions["D"].width = 3.0
            col_letters = "EFGHIJKLMNOPQRSTUVWXYZ"
            for i in range(1, n_ab_ws):
                base_idx = (i - 1) * 2
                if base_idx < len(col_letters):
                    ws.column_dimensions[col_letters[base_idx]].width = 15.0
                if base_idx + 1 < len(col_letters):
                    ws.column_dimensions[col_letters[base_idx + 1]].width = 9.0

    # ── Aba "Tab" — base completa ──────────────────────────────────────────────
    ws_tab = wb.create_sheet("Tab")
    _escrever_sheet(ws_tab, df, ab_list, total_respostas)

    # ── Abas de filtro (uma por valor único de cada coluna em filtro_cols) ──────
    # Suporta tanto filtro_col (str, legado) quanto filtro_cols (list, novo)
    _filtro_cols: list[str] = []
    if filtro_cols:
        _filtro_cols = [c for c in filtro_cols if c and c in df.columns]
    elif filtro_col and filtro_col in df.columns:
        _filtro_cols = [filtro_col]

    _used_sheet_names: set[str] = {"Tab"}
    for fc in _filtro_cols:
        for val in sorted(df[fc].dropna().astype(str).unique()):
            df_f = df[df[fc].astype(str) == val].reset_index(drop=True)
            if len(df_f) == 0:
                continue
            ab_f = preparar_aberturas(df_f, aberturas_cols or [])
            # Nome da aba: caracteres inválidos no Excel removidos, máx 31 chars
            safe = re.sub(r'[\\/*?\[\]:]', '', str(val)).strip()
            base_name = safe[:31] or f"Aba_{len(_used_sheet_names)}"
            sheet_name = base_name
            suffix = 2
            while sheet_name in _used_sheet_names:
                sheet_name = f"{base_name[:28]}_{suffix}"
                suffix += 1
            _used_sheet_names.add(sheet_name)
            ws_f = wb.create_sheet(sheet_name)
            _escrever_sheet(ws_f, df_f, ab_f, len(df_f))

    # ── Aba Base — dados brutos ────────────────────────────────────────────────
    ws_base = wb.create_sheet("Base")
    for ci, col_name in enumerate(df.columns, 1):
        c = ws_base.cell(row=1, column=ci, value=str(col_name))
        c.font = Font(name="Calibri", bold=True, size=10)
    for ri, (_, row_data) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row_data, 1):
            # openpyxl não aceita pd.NA/NaT — escreve célula vazia
            if val is not None and pd.isna(val):
                val = None
            ws_base.cell(row=ri, column=ci, value=val)

    wb.save(saida)
    return saida
